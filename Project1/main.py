# -*- coding: utf-8 -*-
"""
Created on Tue Sep 27 21:20:30 2016

@author: ANMISHA
"""

import openpyxl
import numpy
#from decimal import Decimal
from scipy import stats
from scipy.stats import norm
import math
#import matplotlib.pyplot as plt
wb = openpyxl.load_workbook('university data.xlsx')
sheet = wb.get_sheet_by_name('university_data')
lst1 = []
lst2 = []
lst3 = []
lst4 = []
pi=3.14
for x in range(2,51):
  t=sheet.cell(row=x,column=3).value
  lst1.append(t)
  t=sheet.cell(row=x,column=4).value
  lst2.append(t)
  t=sheet.cell(row=x,column=5).value
  lst3.append(t)
  t=sheet.cell(row=x,column=6).value
  lst4.append(t)
mu1=numpy.mean(lst1)  
mu2=numpy.mean(lst2) 
mu3=numpy.mean(lst3) 
mu4=numpy.mean(lst4)
var1=numpy.var(lst1)
var2=numpy.var(lst2)
var3=numpy.var(lst3)
var4=numpy.var(lst4)
sigma1=numpy.std(lst1)
sigma2=numpy.std(lst2)
sigma3=numpy.std(lst3)
sigma4=numpy.std(lst4)
mu1=round(mu1,3)
mu2=round(mu2,3)
mu3=round(mu3,3)
mu4=round(mu4,3)
var1=round(var1,3)
var2=round(var2,3)
var3=round(var3,3)
var4=round(var4,3)
sigma1=round(sigma1,3)
sigma2=round(sigma2,3)
sigma3=round(sigma3,3)
sigma4=round(sigma4,3)
print ("UBIT name: anmishar")
print ("personNumber: 50208673")
print ("mu1: = "+str(mu1))
print ("mu2: = "+str(mu2))
print ("mu3: = "+str(mu3))
print ("mu4: = "+str(mu4))
print ("var1 = "+str(var1))
print ("var2 = "+str(var2))
print ("var3 = "+str(var3))
print ("var4 = "+str(var4))
print ("standard deviation = "+str(sigma1))
print ("standard deviation = "+str(sigma2))
print ("standard deviation = "+str(sigma3))
print ("standard deviation = "+str(sigma4))

l1="CS Score"
l2="Research Overhead"
l3=" Admin Base Pay"
l4="Tuition"
#x1=lst1
#y1=lst2
#fig1=plt.figure()
#ax1=fig1.add_subplot(111)
#ax1=plt.scatter(x1,y1)
#plt.xlabel(l1)
#plt.ylabel(l2)
#x2=lst1
#y2=lst3
#fig2=plt.figure()
#ax2=fig2.add_subplot(111)
#ax2=plt.scatter(x2,y2)
#plt.xlabel(l1)
#plt.ylabel(l3)
#x3=lst1
#y3=lst4
#fig3=plt.figure()
#ax3=fig3.add_subplot(111)
#ax3=plt.scatter(x3,y3)
#plt.xlabel(l1)
#plt.ylabel(l4)
#x4=lst2
#y4=lst3
#fig4=plt.figure()
#ax4=fig4.add_subplot(111)
#ax4=plt.scatter(x4,y4)
#plt.xlabel(l2)
#plt.ylabel(l3)
#x5=lst2
#fig5=plt.figure()
#y5=lst4
#ax5=fig5.add_subplot(211)
#ax5=plt.scatter(x5,y5)
#plt.xlabel(l2)
#plt.ylabel(l4)
#x6=lst3
#y6=lst4
#fig6=plt.figure()
#ax6=fig6.add_subplot(111)
#ax6=plt.scatter(x6,y6)
#plt.xlabel(l3)
#plt.ylabel(l4)
#-------------------Cov,Corr---------------------------------------
C1 = (numpy.cov(lst1, lst2)[0][0])
C2 = numpy.cov(lst1, lst2)[1][0]
C3 = numpy.cov(lst1, lst2)[1][1]
C4 = numpy.cov(lst1, lst3)[0][1]
C5 = numpy.cov(lst1, lst3)[1][1]
C6 = numpy.cov(lst1, lst4)[0][1]
C7 = numpy.cov(lst1, lst4)[1][1]
C8 = numpy.cov(lst2, lst3)[0][1]
C9 = numpy.cov(lst2, lst4)[0][1]
C10 = numpy.cov(lst3, lst4)[0][1]
t1=round(C1,3)
#print("C1",t1)
t2=round(C2,3)
#print("C2",t2)
t3=round(C3,3)
t4=round(C4,4)
#print("C4",C4)
t5=round(C5,3)
#print("C5",C5)
t6=round(C6,3)
#print("C6",C6)
t7=round(C7,3)
#print("C7",C7)
t8=round(C8,3)
#print("C8",C8)
t9=round(C9,3)
#print("C9",C9)
t10=round(C10,3)
#print("C10",C10)
#co = [[t1,t2,t4,t6],[t2,t3,t8,t9],[t4,t8,t5,t10],[t6,t9,t10,t7]]
co=[t1,round(C2,3),round(C4,3),round(C6,3)],[round(C2,3),round(C3,3),round(C8,3),round(C9,3)],[round(C4,3),round(C8,3),round(C5,3),round(C10,3)],[round(C6,3),round(C9,3),round(C10,3),round(C7,3)]
#print(co)
covarianceMat = numpy.matrix(co)
#numpy.around(covarianceMat.astype(numpy.double),3)
#numpy.matrix.round(covarianceMat)
#cm=numpy.cov(numpy.vstack((lst1,lst2,lst3,lst4)))
#print "cm",cm
print("covarianceMat = \n" +str(covarianceMat ))
#print numpy.corrcoef(a,b)
C11 = (numpy.corrcoef(lst1, lst2)[0][0])
C12 = numpy.corrcoef(lst1, lst2)[1][0]
C13 = numpy.corrcoef(lst1, lst2)[1][1]
C14 = numpy.corrcoef(lst1, lst3)[0][1]
C15 = numpy.corrcoef(lst1, lst3)[1][1]
C16 = numpy.corrcoef(lst1, lst4)[0][1]
C17 = numpy.corrcoef(lst1, lst4)[1][1]
C18 = numpy.corrcoef(lst2, lst3)[0][1]
C19 = numpy.corrcoef(lst2, lst4)[0][1]
C110 = numpy.corrcoef(lst3, lst4)[0][1]
t11=round(C11,3)
#print("C1",t11)
t12=round(C12,3)
#print("C2",t2)
t13=round(C13,3)
t14=round(C14,4)
#print("C4",C4)
t15=round(C15,3)
#print("C5",C5)
t16=round(C16,3)
#print("C6",C6)
t17=round(C17,3)
#print("C7",C7)
t18=round(C18,3)
#print("C8",C8)
t19=round(C19,3)
#print("C9",C9)
t110=round(C110,3)
#print("C10",C10)
co = [[t11,t12,t14,t16],[t12,t13,t18,t19],[t14,t18,t15,t110],[t16,t19,t110,t17]]
correlationMat= numpy.matrix(co)
#cm1=numpy.corrcoef(numpy.vstack((lst1,lst2,lst3,lst4)))
#print "cm1",cm1
print ("correlationMat:\n"+str(correlationMat))
#------------------Loglikelihood---------------------------------
y=list()
n=0
#def normpdf(x, mu, sigma):
#    u = (x-mu)/(sigma)
#    l = (1/(math.sqrt(2*pi)*sigma))*math.exp(-(u*u)/2)
#    m = math.log(l)
#    return m
#for x in lst1:
#    w=normpdf(x,mu1,sigma1)
#    y.append(w)
#for i1 in y:
#    n+=i
#print n
a=0 
b=0
c=0
d=0
m=0
t1=norm(mu1,sigma1).pdf(lst1)
#print "pdf:",t
for i in t1:
    a+= math.log(i)
#print"f1:",a
#t1=norm(mu1,sigma1).logpdf(lst1)
#a=sum(t1)
t2=norm(mu2,sigma2).pdf(lst2)
#print "pdf:",t2
for j in t2:
    b+= math.log(j)
#print"f2:",b
#t2=norm(mu2,sigma2).logpdf(lst2)
#b=sum(t2)
t3=norm(mu3,sigma3).pdf(lst3)
#print "pdf:",t
for k in t3:
    c+= math.log(k)
#print"f3:",c
#t3=norm(mu3,sigma3).logpdf(lst3)
#c=sum(t3)
t4=norm(mu4,sigma4).pdf(lst4)
#print "pdf:",t
for l in t4:
    d+= math.log(l)
#t4=norm(mu4,sigma4).pdf(lst4)
#d=sum(t4)
#print"f4:",d
#t5=norm(mu4,sigma4).logpdf(lst4)
#d=sum(t5)
#print"f5",m
s=a+b+c+d
s=round(s,3)
print ("logLikelihood:"+str(s))
lltotal=[a,b,c,d]
#-----------------------------Bayesian Networks----------------------
arr1=numpy.array(lst1)
arr2=numpy.array(lst2)
arr3=numpy.array(lst3)
arr4=numpy.array(lst4)
size=arr1.size
ad= numpy.zeros(size)
for n in range(ad.size):
    ad[n]=1.0
md=[ad,arr2,arr3,arr4]
cfmat=numpy.zeros((4,4))
for i in range(4):
    for j in range(4):
        sumv=0
        for k in range(size):
            z=(md[i][k]*md[j][k])
            sumv+=z
        cfmat[i][j]=sumv
ymat=numpy.zeros((4,1))
for n in range(4):
    sumv=0
    for i in range(size):
        x=(md[n][i] * arr1[i])
        sumv += x
        ymat[n] = sumv
bn=[[0,0,0,0],[1,0,0,0],[1,0,0,0],[1,0,0,0]]
BNgraph=numpy.matrix(bn)
print ("BNgraph:\n"+str(BNgraph))

barray=numpy.linalg.solve(cfmat,ymat)


sumsq=0
for k in range(size):
    sum1 = 0
    for i in range(4):
        sum1+=md[i][k]*barray[i][0]
    sum1-=arr1[k]
    sum1=sum1*sum1
    sumsq+=(sum1)
sigma=math.sqrt((sumsq/size))                  
var = (sumsq/size)
#-----------------------------loglikelihood----------------
const=-(0.5*numpy.log(2*numpy.pi*var))
BNlog=0
for n in range(size):
    sum1=0
    BNlog += const
    for i in range(len(md)):
        t= (md[i][n]*barray[i][0])
        sum1=sum1+t
    f=sum1-arr1[n]
    sum1=f
    sum1=sum1*sum1
    t1=(-(sum1)/(2*var))
    b1=BNlog+t1
    BNlog=b1
    n=n+1
for i in range(len(lltotal)):
    if i!=0:
        b2=BNlog
        BNlog = b2+lltotal[i]
print ("BNlogLikelihood:"+str(round(BNlog,3)))